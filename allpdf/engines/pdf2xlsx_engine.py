# allpdf/engines/pdf2xlsx_engine.py
"""PDF to XLSX engine using tabula-py + camelot for table extraction."""
import os
import time
from pathlib import Path

from openpyxl import Workbook

from allpdf.engines.base import ConversionEngine
from allpdf.models import ConversionResult, ConversionStatus, FileFormat


class Pdf2XlsxEngine(ConversionEngine):
    """Convert PDF to XLSX by extracting tables.

    Tries camelot (lattice mode) first for bordered tables, then falls
    back to tabula (stream mode). If no structured tables are detected,
    creates sheets with extracted text per page.
    """

    name = "pdf2xlsx"
    input_format = FileFormat.PDF
    output_format = FileFormat.XLSX

    def convert(self, input_path: str, output_path: str, **options) -> ConversionResult:
        start = time.time()

        if not os.path.exists(input_path):
            return ConversionResult(
                input_path=Path(input_path),
                output_path=Path(output_path),
                input_format=FileFormat.PDF,
                output_format=FileFormat.XLSX,
                status=ConversionStatus.FAILED,
                engine_used=self.name,
                duration_seconds=time.time() - start,
                error_message=f"Input file not found: {input_path}",
            )

        try:
            tables = self._extract_tables(input_path, **options)
            self._write_xlsx(tables, output_path, input_path)

            duration = time.time() - start

            return ConversionResult(
                input_path=Path(input_path),
                output_path=Path(output_path),
                input_format=FileFormat.PDF,
                output_format=FileFormat.XLSX,
                status=ConversionStatus.SUCCESS,
                engine_used=self.name,
                duration_seconds=duration,
            )

        except Exception as e:
            duration = time.time() - start
            return ConversionResult(
                input_path=Path(input_path),
                output_path=Path(output_path),
                input_format=FileFormat.PDF,
                output_format=FileFormat.XLSX,
                status=ConversionStatus.FAILED,
                engine_used=self.name,
                duration_seconds=duration,
                error_message=str(e),
            )

    def _extract_tables(self, input_path: str, **options) -> list[list[list[str]]]:
        """Extract tables from PDF. Tries camelot first, falls back to tabula."""
        tables = []

        try:
            import camelot
            camelot_tables = camelot.read_pdf(input_path, pages="all", flavor="lattice")
            for t in camelot_tables:
                tables.append(t.df.values.tolist())
        except Exception:
            pass

        if not tables:
            try:
                import tabula
                tabula_tables = tabula.read_pdf(input_path, pages="all", multiple_tables=True)
                for df in tabula_tables:
                    if not df.empty:
                        tables.append(df.values.tolist())
            except Exception:
                pass

        return tables

    def _write_xlsx(self, tables: list[list[list[str]]], output_path: str, input_path: str):
        """Write extracted tables to an XLSX workbook."""
        wb = Workbook()
        wb.remove(wb.active)

        if tables:
            for i, table in enumerate(tables):
                ws = wb.create_sheet(title=f"Table_{i+1}")
                for row_idx, row in enumerate(table, start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=str(cell) if cell is not None else "")
        else:
            import fitz
            doc = fitz.open(input_path)
            for page_num, page in enumerate(doc):
                text = page.get_text()
                if text.strip():
                    ws = wb.create_sheet(title=f"Page_{page_num+1}")
                    for row_idx, line in enumerate(text.strip().split("\n"), start=1):
                        ws.cell(row=row_idx, column=1, value=line)
            doc.close()

            if not wb.sheetnames:
                ws = wb.create_sheet(title="Sheet1")
                ws.cell(row=1, column=1, value="No extractable content found")

        wb.save(output_path)

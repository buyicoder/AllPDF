# allpdf/utils.py
import os
import tempfile
from pathlib import Path

from allpdf.models import FileFormat


def temp_output_path(input_path: str, target_format: str) -> str:
    """Generate a temporary output path alongside the input file."""
    in_path = Path(input_path)
    stem = in_path.stem
    out_dir = in_path.parent
    out_path = out_dir / f"{stem}_converted.{target_format}"
    return str(out_path)


def detect_format(path: Path) -> FileFormat:
    """Detect file format from extension."""
    return FileFormat.from_path(path)


def ensure_dir(path: str) -> Path:
    """Ensure a directory exists, creating it if needed."""
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def count_pdf_pages(filepath: str) -> int:
    """Count pages in a PDF file using PyMuPDF."""
    import fitz
    doc = fitz.open(filepath)
    count = doc.page_count
    doc.close()
    return count


def count_pdf_images(filepath: str) -> int:
    """Count embedded images in a PDF file."""
    import fitz
    doc = fitz.open(filepath)
    total = 0
    for page in doc:
        total += len(page.get_images())
    doc.close()
    return total


def count_docx_pages(filepath: str) -> int:
    """Count pages in a DOCX file (approximate via page breaks + content)."""
    return _count_office_pages_via_libreoffice(filepath, "docx")


def count_xlsx_sheets(filepath: str) -> int:
    """Count sheets in an XLSX file."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    count = len(wb.sheetnames)
    wb.close()
    return count


def count_pptx_slides(filepath: str) -> int:
    """Count slides in a PPTX file."""
    from pptx import Presentation
    prs = Presentation(filepath)
    return len(prs.slides)


def _count_office_pages_via_libreoffice(filepath: str, fmt: str) -> int:
    """Use LibreOffice to convert to PDF temporarily and count pages."""
    import subprocess
    with tempfile.TemporaryDirectory() as tmpdir:
        result = subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf", "--outdir", tmpdir, filepath],
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode != 0:
            return 0
        pdf_path = os.path.join(tmpdir, Path(filepath).stem + ".pdf")
        if os.path.exists(pdf_path):
            return count_pdf_pages(pdf_path)
    return 0
